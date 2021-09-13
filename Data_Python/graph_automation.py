import time
from typing import List
import pandas as pd
import glob
#변경할 파일들이 모여있는 경로
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
start = time.time()
#  전처리할 파일이 모여있는 경로
file_address_local="C:/Users/20615/Desktop/python code/Python_Project/Raw file _ to IT/"
#파일 Union, create new xlsx, 파일 번호까지
def file_union(local_file_address, file_num):
    file_ad = []
    dataframe_list=[]

    for i in range(len(file_get_file_address(local_file_address)[0])):
        dff = []
        for file in file_get_file_address(local_file_address)[0][i]:
            csv_df = pd.read_csv(file)
            csv_df = csv_df.drop(['DUT'],axis=1)
            csv_df = csv_df[csv_df.columns[:-1]]
            dff.append(csv_df)

        result3 = pd.concat(dff, axis=1)
        dataframe_list.append(result3)
        #파일저장
        file_ad.append("[" + str(file_num[i]) + "]" + ".xlsx")
    return file_ad, dataframe_list


#파일의 정보를 읽어옴
def file_get_file_address(file_address):
    address1 = []
    cf = []
    file_num = []
    file_name = []
    for i in range(20):
        cf.append(i+1)
    for k in cf:
        address = []
        for file in glob.glob(file_address + "[" + str(k) + "*"):
            address.append(file)
            continue
        if len(address) != 0:
            address1.append(address)
            file_num.append(k)
            file_name.append(file.split('\\')[1].split(".CSV")[0].split('] ')[1])
    return address1, file_num

#파일의 경로까지 읽어야함 ~~.csv까지
def data_processing(dataframe, union_file_num):
    df=dataframe
    df = df.where(pd.notnull(df), "")
    dc=list(df.columns)
    #첫 번째 행 데이터 저장
    second_df=df.loc[0]
    sd=[]
    for i in range(len(second_df)):
        sd.append(second_df[i])
    # 첫번쨰 행 제거
    df = df.drop(index=0, axis=0)
    # time columns 조회 후 데이터 변환
    dt=[]
    for x in df.columns[0::2]:
        dt.append(x)
    time=[]
    for i in range(len(dt)):
        time.append(list(df[dt[i]]))
    #Hour 만 추출
    time2=[]
    for k in range(len(dt)):
        time1 = []
        for i in range(len(time[k])):
            time[k][i]=time[k][i].strip()
            time[k][i]=time[k][i].split(":")[0]
            if time[k][i] != "":
                time[k][i]=int(time[k][i])
            time1.append(time[k][i])
        time2.append(time1)

    for i in range(len(dt)):
        df[dt[i]]=time2[i]
    df.apply(pd.to_numeric)

    for i in dc:
        df[i]=pd.to_numeric(df[i])
    # 지웠던 sd 첫번째 행 삽입
    new_data = dict(zip(dc, sd))
    temp1 = df[df.index < 0]
    temp2 = df[df.index >= 0]
    df = temp1.append(new_data,ignore_index=True).append(temp2, ignore_index=True)
    #엑셀화
    df.to_excel(excel_writer=str(union_file_num) + "_union" +".xlsx", index=False)
    file_name=str(union_file_num) + "_union" +".xlsx"
    return file_name
# 데이터 엑셀 생성 함수
def data_creation(file_address_local) -> List:
    # Union 데이터 생성
    dataframe_list = file_union(file_address_local, file_get_file_address(file_address_local)[1])[1]
    # 파일 숫자
    file_n = file_get_file_address(file_address_local)[1]
    processed_file_name=[]
    # dataframe list input
    for i in range(len(dataframe_list)):
        processed_file_name.append(data_processing(dataframe_list[i], file_n[i]))
    return processed_file_name

for d in data_creation(file_address_local):
    wb=load_workbook(d)
    ws= wb.active
    max_r = ws.max_row
    max_c = ws.max_column
    chart = ScatterChart()
    xvalues = Reference(ws, min_row=1, max_row= max_r, min_col = 1)
    values = Reference(ws, min_row = 1, max_row= max_r, min_col = 2)
    series = Series(values, xvalues, title_from_data= True)
    chart.series.append(series)
    col = (max_c)//2
    chart.x_axis.title = 'Time_count'
    chart.y_axis.title = 'Delta R(%)'
    chart.type = "col"
    for j in range(1,col):
        chart2 = ScatterChart()
        xvalues = Reference(ws, min_row=1, max_row=max_r, min_col=1 + (2 * j))
        values = Reference(ws, min_row=1, max_row=max_r, min_col=2 + (2 * j))
        series = Series(values, xvalues, title_from_data=True)
        chart2.series.append(series)
        chart2.style = 13
        chart2.x_axis.title = 'Time_count'
        chart2.y_axis.title = 'Delta R(%)'
        chart += chart2
    ws.add_chart(chart, "F5")
    wb.save("graph_" + d)
    print("graph_" + d+" DONE")

print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간