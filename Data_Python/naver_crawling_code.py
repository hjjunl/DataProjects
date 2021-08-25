import os
import time
from datetime import datetime

import openpyxl
import options as options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager  # 'webdriver_manager' 패키지모듈 다운로드 필요
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from typing import List, Tuple, Dict, Final
import pandas as pd
import fsspec
from termcolor import colored
from datetime import datetime, timedelta

options = Options()
WAIT_TIME: Final = 2

url1 = 'https://search.naver.com/search.naver?where=news&sm=tab_jum&query='
# searching_word = input()
searching_word = '코로나'
url = url1 + searching_word
driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)  # 최신 크롬 드라이버 설치
EXCEL_FILE = "Naver_Data.xlsx"


def save(lv_df):
    lv_wb = load_workbook(EXCEL_FILE)
    lv_ws = lv_wb['Sheet1']
    lv_last_row = lv_ws.max_row
    for lv_i, lv_row in lv_df.iterrows():
        lv_ws.cell(lv_last_row + lv_i + 1, 1).value = lv_row['뉴스제목']
        lv_ws.cell(lv_last_row + lv_i + 1, 2).value = lv_row["신문사"]
        lv_ws.cell(lv_last_row + lv_i + 1, 3).value = f'=HYPERLINK("{lv_row["뉴스 링크"]}","링크 접속")'
        lv_ws.cell(lv_last_row + lv_i + 1, 4).value = lv_row['뉴스 날짜']
        lv_ws.cell(lv_last_row + lv_i + 1, 5).value = lv_row['줄거리']

    lv_wb.save(EXCEL_FILE)
    lv_wb.close()


def wait_element_ready(_driver: webdriver, xpath: str, wait: int = WAIT_TIME) -> WebElement:
    WebDriverWait(_driver, wait).until(
        expected_conditions.presence_of_all_elements_located(
            (By.XPATH, xpath)))
    web_element = _driver.find_element_by_xpath(xpath)
    return web_element


def wait_elements_ready(_driver: webdriver, xpath: str, wait: int = WAIT_TIME) -> List:
    WebDriverWait(_driver, wait).until(
        expected_conditions.presence_of_all_elements_located(
            (By.XPATH, xpath)))
    web_elements = _driver.find_elements_by_xpath(xpath)
    return web_elements


# 검색어 치기-> '옵션' 선택 -> '직접 입력' 클릭 -> 년 월 일 변경후 크롤링 다한후 없으면 다시 반복->
# 몇 페이지 가져올지 정하기 input
# driver.get(url)
# wait_element_ready(driver, "//input[@name='lot']").send_keys(Keys.ENTER)
news_title_list = []
news_overview_list = []
news_date_list = []
news_press_list = []
news_link_list = []
change_date_type = []
news_press_list_change = []


def main():
    # 시작날짜 입력
    # start_date1 = input()  # 예 20210804
    start_date1 = '20210404'
    start_date = datetime.strptime(start_date1, '%Y%m%d')
    # 말일 입력
    # end_date1 = input()
    end_date1 = '20210406'
    end_date = datetime.strptime(end_date1, '%Y%m%d')  # 날짜를 입력할 리스트
    str_date_list = []
    while start_date.strftime('%Y%m%d') != end_date.strftime('%Y%m%d'):
        str_date_list.append(start_date.strftime('%Y%m%d'))
        start_date += timedelta(days=1)
    str_date_list = list(map(int, str_date_list))

    for dat in str_date_list:
        driver.get(url)
        start_year = dat // 10000
        start_month = (dat % 10000) // 100
        start_day = (dat) % 100
        # 날짜 오픈
        try:
            wait_element_ready(driver, "//*[@id='snb']/div[1]/div/div[2]/a").send_keys(Keys.ENTER)  # ""는 ''로 바꿔줌
            # driver.find_element_by_class_name('_search_option_open_btn').click()
            wait_element_ready(driver, "//*[@id='snb']/div[1]/div/div[2]/a").send_keys(Keys.ENTER)  # ""는 ''로 바꿔줌
            wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[1]/a[9]").click()  # ""는 ''로
        except:
            continue
        # 시작일 클릭
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[1]/div/div/div/ul/li[" + str(
            start_year - 1989) + "]/a").click()  # 년도
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[2]/div/div/div/ul/li[" + str(
            start_month) + "]/a").click()  # 월
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[3]/div/div/div/ul/li[" + str(
            start_day) + "]/a").click()  # 일
        # 디시 시작일 클릭
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[1]/span[3]/a").click()  # 직접입력
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[1]/div/div/div/ul/li[" + str(
            start_year - 1989) + "]/a").click()  # 년도
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[2]/div/div/div/ul/li[" + str(
            start_month) + "]/a").click()  # 월
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[2]/div[3]/div/div/div/ul/li[" + str(
            start_day) + "]/a").click()  # 일

        # 적용 버튼
        wait_element_ready(driver, "//*[@id='snb']/div[2]/ul/li[2]/div/div[3]/div[3]/button").click()
        # news_title=wait_elements_ready(driver, "")
        system_date_format = '%Y.%m.%d'

        for page in range(1, 3):
            # 뉴스 제목 크롤링
            news_title = driver.find_elements_by_class_name("news_tit")

            # 뉴스 간단 정보 크롤링
            news_overview = driver.find_elements_by_class_name("api_txt_lines.dsc_txt_wrap")

            # 신문사 클롤링
            news_press = driver.find_elements_by_class_name("info.press")

            for i in news_overview:
                # print(i.text)
                news_overview_list.append(i.text)
            for i in range(len(news_overview)):
                news_date_list.append(dat)
            for i in news_title:
                # 뉴스 링크
                a = i.get_attribute('href')
                news_title_list.append(i.text)
                news_link_list.append(a)
            for i in news_press:
                # print(i.text)
                news_press_list.append(i.text)

            # 페이지가 존재하지 않을때 그냥 진행
            try:
                wait_element_ready(driver, "//*[@id='main_pack']/div[2]/div/div/a[" + str(page) + "]").click()  # 페이지 수
            except:
                pass
    # 신문사 불필요 단어 제거
    for i in range(len(news_press_list)):
        if '언론사 선정' in news_press_list[i]:
            temp = news_press_list[i].replace('언론사 선정', '')
            news_press_list_change.append(temp)
        else:
            news_press_list_change.append(i)
    # 날짜 포멧
    for i in news_date_list:
        change_date_type.append(pd.to_datetime(str(i), format='%Y%m%d').strftime("%Y-%m-%d"))
    # 엑셀생성
    dd = {'뉴스제목': news_title_list, '신문사': news_press_list_change, '뉴스 링크': news_link_list,
          '뉴스 날짜': change_date_type, '줄거리': news_overview_list}
    # '뉴스 날짜': news_date_list
    df = pd.DataFrame(dd)
    df.set_index('뉴스제목', inplace=True)
    if os.path.isfile("Naver_Data.xlsx"):
        df = df.reset_index()
        save(df)
    else:
        df.to_excel(excel_writer='Naver_Data.xlsx')
    driver.close()


# 엑셀에서 연합뉴스 색 입히기, 테두리
def color_keyword(excel_file):
    excel_document = openpyxl.load_workbook(excel_file)
    ws = excel_document.active
    excel_document.get_sheet_names()
    lv_ws = excel_document['Sheet1']
    lv_last_row = lv_ws.max_row
    lv_last_column = lv_ws.max_column
    multiple_cells = lv_ws['B2':'B' + str(lv_last_row)]
    count = 2
    arr = []
    for row in multiple_cells:
        for cell in row:
            if '연합뉴스' in str(cell.value):
                arr.append(count)
                print(count)
            count += 1
    for j in arr:
        ws.cell(row=j, column=2).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    THICK_BORDER = Border(Side('thick'), Side('thick'), Side('thick'), Side('thick'))
    #범위 설정시 라인이생김
    for rng in ws['J2:K10']:
        for cell in rng:
            cell.border = THICK_BORDER  # [J2:K10] 모든테두리 설정
    excel_document.save(filename=EXCEL_FILE)


main()
color_keyword(EXCEL_FILE)
